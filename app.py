from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from twilio.rest import Client
from twilio.base.exceptions import TwilioRestException
import mysql.connector
from dotenv import load_dotenv
from functools import wraps
from datetime import date, timedelta
from io import BytesIO
import re
import os
import json
import logging
load_dotenv()

app = Flask(__name__)
app.secret_key = "sgt_modulo3"

logging.basicConfig(level=logging.INFO)

db_config = {
    'host': os.getenv('DB_HOST'),
    'database': os.getenv('DB_DATABASE'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
}


def get_db_connection():
    return mysql.connector.connect(**db_config)


# ---------------------------------------------------------
# GARANTE A TABELA DE NOTIFICAÇÕES INDIVIDUAIS (idempotente)
# ---------------------------------------------------------
def garantir_tabela_notificacoes():
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS notificacoes (
                ID_NOTIF INT AUTO_INCREMENT PRIMARY KEY,
                ID_FUN INT NOT NULL,
                ID_ADM INT NULL,
                TITULO VARCHAR(150) NOT NULL,
                MENSAGEM TEXT NOT NULL,
                TIPO_ENVIO ENUM('geral', 'individual') NOT NULL DEFAULT 'individual',
                DATA_ENVIO DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                LIDA TINYINT(1) NOT NULL DEFAULT 0,
                CONSTRAINT fk_notif_fun FOREIGN KEY (ID_FUN) REFERENCES funcionarios(ID_FUN) ON DELETE CASCADE
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
    except Exception:
        app.logger.exception("Não foi possível garantir a tabela de notificações")
        if conn and conn.is_connected():
            conn.close()


garantir_tabela_notificacoes()


# ---------------------------------------------------------
# GARANTE A TABELA DE LIXEIRA (idempotente)
# Guarda uma "foto" (JSON) de qualquer registro excluído em
# funcionarios/adm/treinamentos/relatorio, permitindo restaurar depois.
# ---------------------------------------------------------
def garantir_tabela_lixeira():
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS lixeira (
                ID_LIXEIRA INT AUTO_INCREMENT PRIMARY KEY,
                TABELA_ORIGEM VARCHAR(50) NOT NULL,
                ID_ORIGINAL INT NOT NULL,
                DADOS JSON NOT NULL,
                EXCLUIDO_POR VARCHAR(150) DEFAULT NULL,
                DATA_EXCLUSAO DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
    except Exception:
        app.logger.exception("Não foi possível garantir a tabela de lixeira")
        if conn and conn.is_connected():
            conn.close()


garantir_tabela_lixeira()


# Colunas restauráveis por tabela de origem (usadas tanto para montar o
# INSERT de restauração quanto para validar o que pode ser restaurado).
TABELAS_RESTAURAVEIS = {
    'funcionarios': ['ID_FUN', 'NOME_FUN', 'EMAIL_FUN', 'CPF_FUN', 'TELEFONE_FUN', 'CARGO_FUN', 'SETOR_FUN', 'WHATSAPP_FUN', 'SENHA_FUN'],
    'adm': ['ID_ADM', 'NOME_ADM', 'EMAIL_ADM', 'CPF_ADM', 'TELEFONE_ADM', 'CARGO_ADM', 'SETOR_ADM', 'WHATSAPP_ADM', 'SENHA_ADM'],
    'treinamentos': ['ID_TREIN', 'NOME_NR', 'VALIDADE_DIAS', 'DESCRICAO_NR'],
    'relatorio': ['ID_REL', 'ID_FUN', 'ID_TREIN', 'DATA_REALIZACAO', 'DATA_VENCIMENTO'],
}


def _serializador_json(obj):
    """Converte tipos que o json padrão não entende (datas) para texto."""
    if isinstance(obj, date):
        return obj.isoformat()
    return str(obj)


def mover_para_lixeira(cursor, tabela, id_original, dados_registro, excluido_por):
    """
    Grava uma cópia (JSON) do registro que está prestes a ser excluído na
    tabela `lixeira`, para permitir recuperação posterior. Deve ser chamada
    ANTES do DELETE, usando a mesma conexão/transação (só é persistido de
    fato quando o conn.commit() da rota for chamado).
    """
    cursor.execute(
        "INSERT INTO lixeira (TABELA_ORIGEM, ID_ORIGINAL, DADOS, EXCLUIDO_POR) VALUES (%s, %s, %s, %s)",
        (tabela, id_original, json.dumps(dados_registro, default=_serializador_json, ensure_ascii=False), excluido_por)
    )


# ---------------------------------------------------------
# FUNÇÃO AUXILIAR: Sanitização e Formatação E.164
# ---------------------------------------------------------
def formatar_para_twilio(telefone_raw):
    if not telefone_raw:
        return ""
    numeros = re.sub(r'\D', '', str(telefone_raw))
    if len(numeros) in (10, 11):
        return f"+55{numeros}"
    elif len(numeros) in (12, 13):
        return f"+{numeros}"
    return numeros


# ---------------------------------------------------------
# FUNÇÃO AUXILIAR: Status do treinamento a partir do vencimento
# (mesma regra de 30 dias usada em /api/funcionario/realizar-treinamento
# e espelhada no front-end em static/js/status-utils.js)
# ---------------------------------------------------------
def status_relatorio(data_vencimento):
    if not data_vencimento:
        return ''
    hoje = date.today()
    limite_30_dias = hoje + timedelta(days=30)
    if data_vencimento < hoje:
        return 'vencido'
    if data_vencimento <= limite_30_dias:
        return 'vencendo'
    return 'valido'


# =========================================================
# AUTENTICAÇÃO / SESSÃO
# =========================================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_type' not in session:
            if request.path.startswith('/api/'):
                return jsonify({"erro": "Não autenticado"}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def adm_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('user_type') != 'adm':
            if request.path.startswith('/api/'):
                return jsonify({"erro": "Acesso restrito a administradores"}), 403
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


@app.route('/login', methods=['GET'])
def login_page():
    if 'user_type' in session:
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        dados = request.get_json()
        email = (dados.get('email') or '').strip().lower()
        senha = dados.get('senha') or ''

        if not email or not senha:
            return jsonify({"erro": "Informe e-mail e senha"}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Tenta primeiro como ADM
        cursor.execute("SELECT * FROM adm WHERE LOWER(EMAIL_ADM) = %s", (email,))
        usuario = cursor.fetchone()
        tipo = 'adm'

        if not usuario:
            cursor.execute("SELECT * FROM funcionarios WHERE LOWER(EMAIL_FUN) = %s", (email,))
            usuario = cursor.fetchone()
            tipo = 'funcionario'

        cursor.close()
        conn.close()

        if not usuario:
            return jsonify({"erro": "E-mail ou senha inválidos"}), 401

        senha_hash = usuario['SENHA_ADM'] if tipo == 'adm' else usuario['SENHA_FUN']

        if not senha_hash:
            return jsonify({"erro": "Este usuário ainda não possui senha cadastrada. Contate o administrador."}), 401

        try:
            senha_valida = check_password_hash(senha_hash, senha)
        except Exception:
            senha_valida = False

        if not senha_valida:
            return jsonify({"erro": "E-mail ou senha inválidos"}), 401

        session['user_type'] = tipo
        session['user_id'] = usuario['ID_ADM'] if tipo == 'adm' else usuario['ID_FUN']
        session['user_name'] = usuario['NOME_ADM'] if tipo == 'adm' else usuario['NOME_FUN']

        return jsonify({
            "mensagem": "Login realizado com sucesso!",
            "tipo": tipo,
            "nome": session['user_name'],
            "redirect": url_for('index')
        }), 200

    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))


@app.route('/api/me', methods=['GET'])
@login_required
def api_me():
    return jsonify({
        "tipo": session.get('user_type'),
        "id": session.get('user_id'),
        "nome": session.get('user_name')
    }), 200


# =========================================================
# CONFIGURAÇÕES - DADOS DA PRÓPRIA CONTA (ADM ou FUNCIONÁRIO)
# =========================================================

@app.route('/api/minha-conta', methods=['GET'])
@login_required
def obter_minha_conta():
    try:
        tipo = session.get('user_type')
        user_id = session.get('user_id')
        tabela = 'adm' if tipo == 'adm' else 'funcionarios'
        sufixo = 'ADM' if tipo == 'adm' else 'FUN'

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            f"""SELECT ID_{sufixo} AS ID, NOME_{sufixo} AS NOME, EMAIL_{sufixo} AS EMAIL,
                       TELEFONE_{sufixo} AS TELEFONE, WHATSAPP_{sufixo} AS WHATSAPP,
                       CARGO_{sufixo} AS CARGO, SETOR_{sufixo} AS SETOR
                FROM {tabela} WHERE ID_{sufixo} = %s""",
            (user_id,)
        )
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()

        if not resultado:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        resultado['tipo'] = tipo
        return jsonify(resultado), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/minha-conta', methods=['PUT'])
@login_required
def atualizar_minha_conta():
    conn = None
    try:
        dados = request.get_json(force=True)
        tipo = session.get('user_type')
        user_id = session.get('user_id')
        tabela = 'adm' if tipo == 'adm' else 'funcionarios'
        sufixo = 'ADM' if tipo == 'adm' else 'FUN'

        senha_atual = dados.get('senha_atual') or ''
        if not senha_atual:
            return jsonify({"erro": "Informe sua senha atual para confirmar as alterações"}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"SELECT * FROM {tabela} WHERE ID_{sufixo} = %s", (user_id,))
        existente = cursor.fetchone()

        if not existente:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Usuário não encontrado"}), 404

        senha_hash_atual = existente.get(f'SENHA_{sufixo}')
        try:
            senha_valida = senha_hash_atual and check_password_hash(senha_hash_atual, senha_atual)
        except Exception:
            senha_valida = False

        if not senha_valida:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Senha atual incorreta"}), 401

        senha_nova = dados.get('senha_nova') or ''
        if senha_nova and len(senha_nova) < 6:
            cursor.close()
            conn.close()
            return jsonify({"erro": "A nova senha deve ter pelo menos 6 caracteres"}), 400

        senha_final = generate_password_hash(senha_nova) if senha_nova else senha_hash_atual
        novo_nome = dados.get('nome', existente[f'NOME_{sufixo}'])
        novo_email = (dados.get('email') or existente[f'EMAIL_{sufixo}']).strip().lower()

        cursor.execute(f"""
            UPDATE {tabela}
            SET NOME_{sufixo}=%s, EMAIL_{sufixo}=%s, TELEFONE_{sufixo}=%s, WHATSAPP_{sufixo}=%s, SENHA_{sufixo}=%s
            WHERE ID_{sufixo}=%s
        """, (
            novo_nome,
            novo_email,
            dados.get('telefone', existente[f'TELEFONE_{sufixo}']),
            dados.get('whatsapp', existente[f'WHATSAPP_{sufixo}']),
            senha_final,
            user_id
        ))
        conn.commit()
        cursor.close()
        conn.close()

        session['user_name'] = novo_nome

        return jsonify({"mensagem": "Dados atualizados com sucesso!"}), 200
    except mysql.connector.errors.IntegrityError:
        if conn and conn.is_connected():
            conn.close()
        return jsonify({"erro": "Este e-mail ou telefone já está em uso por outro cadastro."}), 409
    except Exception as e:
        if conn and conn.is_connected():
            conn.close()
        return jsonify({"erro": str(e)}), 500


# =========================================================
# PÁGINAS (RENDERIZAÇÃO)
# =========================================================

@app.route('/')
@login_required
def index():
    return render_template('index.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='dashboard')


@app.route('/funcionarios')
@adm_required
def pagina_funcionarios():
    return render_template('funcionarios.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='funcionarios')


@app.route('/treinamentos')
@login_required
def pagina_treinamentos():
    return render_template('treinamentos.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='treinamentos')


@app.route('/relatorios')
@login_required
def pagina_relatorios():
    return render_template('relatorios.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='relatorios')


@app.route('/notificacoes')
@adm_required
def pagina_notificacoes():
    return render_template('notificacoes.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='notificacoes')


@app.route('/configuracoes')
@login_required
def pagina_configuracoes():
    return render_template('configuracoes.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='configuracoes')


@app.route('/administradores')
@adm_required
def pagina_administradores():
    return render_template('administradores.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='administradores')


@app.route('/lixeira')
@adm_required
def pagina_lixeira():
    return render_template('lixeira.html', tipo=session.get('user_type'), nome=session.get('user_name'), ativo='lixeira')


# =========================================================
# ROTA AUTOMATIZADA: NOTIFICAÇÕES (SMS + WHATSAPP)
# =========================================================

@app.route('/api/notificar-vencimentos', methods=['POST'])
@adm_required
def notificar_vencimentos():
    conn = None
    try:
        account_sid = os.getenv('TWILIO_ACCOUNT_SID')
        auth_token = os.getenv('TWILIO_AUTH_TOKEN')
        sms_from = os.getenv('TWILIO_PHONE_NUMBER')
        whatsapp_from = os.getenv('TWILIO_WHATSAPP_NUMBER')

        client = Client(account_sid, auth_token)

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        sql = """
            SELECT F.NOME_FUN, F.TELEFONE_FUN, T.NOME_NR, R.DATA_VENCIMENTO
            FROM relatorio R
            JOIN funcionarios F ON R.ID_FUN = F.ID_FUN
            JOIN treinamentos T ON R.ID_TREIN = T.ID_TREIN
            JOIN (
                SELECT ID_FUN, ID_TREIN, MAX(DATA_VENCIMENTO) AS ULTIMO_VENCIMENTO
                FROM relatorio
                GROUP BY ID_FUN, ID_TREIN
            ) U ON U.ID_FUN = R.ID_FUN AND U.ID_TREIN = R.ID_TREIN AND U.ULTIMO_VENCIMENTO = R.DATA_VENCIMENTO
            WHERE R.DATA_VENCIMENTO IS NOT NULL
              AND R.DATA_VENCIMENTO <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)
        """
        cursor.execute(sql)
        vencendo = cursor.fetchall()

        contador_sucesso = 0
        for reg in vencendo:
            telefone_limpo = formatar_para_twilio(reg['TELEFONE_FUN'])

            if not telefone_limpo or len(telefone_limpo) < 12:
                print(f"Aviso: Registro de {reg['NOME_FUN']} ignorado por inconsistência no telefone.")
                continue

            venceu = reg['DATA_VENCIMENTO'] < date.today()
            vencimento_br = reg['DATA_VENCIMENTO'].strftime('%d/%m/%Y') if reg['DATA_VENCIMENTO'] else ''

            if venceu:
                mensagem = f"Olá {reg['NOME_FUN']}, seu treinamento de {reg['NOME_NR']} está VENCIDO desde {vencimento_br}. Regularize o quanto antes."
            else:
                mensagem = f"Olá {reg['NOME_FUN']}, seu treinamento de {reg['NOME_NR']} vence em breve no dia {vencimento_br}."

            client.messages.create(body=mensagem, from_=sms_from, to=telefone_limpo)
            client.messages.create(body=mensagem, from_=f"whatsapp:{whatsapp_from}", to=f"whatsapp:{telefone_limpo}")

            contador_sucesso += 1

        return jsonify({"mensagem": f"Processamento concluído. {contador_sucesso} colaboradores notificados!"}), 200

    except TwilioRestException as e:
        # A Twilio devolve erros técnicos (ex.: "Unable to create record: ...
        # [HTTP 400]", "Error 21211: ..."), que não fazem sentido para quem
        # está usando o sistema. Registramos o erro real nos logs do servidor
        # (para investigação) e devolvemos uma mensagem de validação legível
        # ao usuário, em vez do código bizarro do Twilio.
        app.logger.exception("Erro do Twilio ao disparar notificações")
        return jsonify({"erro": "Não foi possível enviar as notificações. Verifique se os números de telefone dos colaboradores e as configurações do Twilio estão corretos e tente novamente."}), 500
    except Exception as e:
        app.logger.exception("Erro crítico no processamento das notificações")
        return jsonify({"erro": "Não foi possível concluir o envio das notificações agora. Tente novamente em instantes."}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


# =========================================================
# NOTIFICAÇÕES INDIVIDUAIS (SISTEMA INTERNO / SINO)
# O ADM pode notificar todos os funcionários (geral) ou escolher
# um funcionário específico (individual). Cada notificação vira
# um registro por funcionário, o que permite marcar como lida de
# forma independente para cada um.
# =========================================================

@app.route('/api/notificacoes', methods=['POST'])
@adm_required
def enviar_notificacao():
    conn = None
    try:
        dados = request.get_json(force=True)
        titulo = (dados.get('titulo') or '').strip()
        mensagem = (dados.get('mensagem') or '').strip()
        destino = dados.get('destino')
        id_fun = dados.get('id_fun')

        if not titulo or not mensagem:
            return jsonify({"erro": "Preencha o título e a mensagem da notificação"}), 400
        if destino not in ('geral', 'individual'):
            return jsonify({"erro": "Escolha o destinatário da notificação"}), 400
        if destino == 'individual' and not id_fun:
            return jsonify({"erro": "Selecione o funcionário que deve receber a notificação"}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        if destino == 'geral':
            cursor.execute("SELECT ID_FUN FROM funcionarios")
            destinatarios = [row['ID_FUN'] for row in cursor.fetchall()]
            if not destinatarios:
                cursor.close()
                conn.close()
                return jsonify({"erro": "Não há funcionários cadastrados para notificar"}), 400
        else:
            cursor.execute("SELECT ID_FUN FROM funcionarios WHERE ID_FUN = %s", (id_fun,))
            if not cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({"erro": "Funcionário não encontrado"}), 404
            destinatarios = [id_fun]

        cursor_escrita = conn.cursor()
        for fid in destinatarios:
            cursor_escrita.execute("""
                INSERT INTO notificacoes (ID_FUN, ID_ADM, TITULO, MENSAGEM, TIPO_ENVIO)
                VALUES (%s, %s, %s, %s, %s)
            """, (fid, session.get('user_id'), titulo, mensagem, destino))
        conn.commit()
        cursor.close()
        cursor_escrita.close()
        conn.close()

        qtd = len(destinatarios)
        plural = "funcionário" if qtd == 1 else "funcionários"
        return jsonify({"mensagem": f"Notificação enviada para {qtd} {plural}!"}), 201
    except Exception:
        if conn and conn.is_connected():
            conn.close()
        app.logger.exception("Erro ao enviar notificação individual")
        return jsonify({"erro": "Não foi possível enviar a notificação agora. Tente novamente em instantes."}), 500


@app.route('/api/notificacoes', methods=['GET'])
@adm_required
def listar_notificacoes_enviadas():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT n.ID_NOTIF, n.TITULO, n.MENSAGEM, n.TIPO_ENVIO, n.DATA_ENVIO, n.LIDA, f.NOME_FUN
            FROM notificacoes n
            LEFT JOIN funcionarios f ON n.ID_FUN = f.ID_FUN
            ORDER BY n.DATA_ENVIO DESC
            LIMIT 200
        """)
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()
        for r in resultados:
            if r.get('DATA_ENVIO'):
                r['DATA_ENVIO'] = r['DATA_ENVIO'].isoformat()
            r['LIDA'] = bool(r['LIDA'])
        return jsonify(resultados), 200
    except Exception:
        app.logger.exception("Erro ao listar notificações enviadas")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/notificacoes/minhas', methods=['GET'])
@login_required
def minhas_notificacoes():
    try:
        if session.get('user_type') != 'funcionario':
            return jsonify([]), 200

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT ID_NOTIF, TITULO, MENSAGEM, DATA_ENVIO, LIDA
            FROM notificacoes
            WHERE ID_FUN = %s
            ORDER BY DATA_ENVIO DESC
            LIMIT 50
        """, (session.get('user_id'),))
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()
        for r in resultados:
            if r.get('DATA_ENVIO'):
                r['DATA_ENVIO'] = r['DATA_ENVIO'].isoformat()
            r['LIDA'] = bool(r['LIDA'])
        return jsonify(resultados), 200
    except Exception:
        app.logger.exception("Erro ao buscar notificações do funcionário")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/notificacoes/<int:id_notif>/lida', methods=['PUT'])
@login_required
def marcar_notificacao_lida(id_notif):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT ID_FUN FROM notificacoes WHERE ID_NOTIF = %s", (id_notif,))
        registro = cursor.fetchone()
        if not registro:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Notificação não encontrada"}), 404

        if session.get('user_type') == 'funcionario' and registro['ID_FUN'] != session.get('user_id'):
            cursor.close()
            conn.close()
            return jsonify({"erro": "Acesso negado"}), 403

        cursor_escrita = conn.cursor()
        cursor_escrita.execute("UPDATE notificacoes SET LIDA = 1 WHERE ID_NOTIF = %s", (id_notif,))
        conn.commit()
        cursor.close()
        cursor_escrita.close()
        conn.close()
        return jsonify({"mensagem": "Notificação marcada como lida"}), 200
    except Exception:
        app.logger.exception("Erro ao marcar notificação como lida")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/notificacoes/marcar-todas-lidas', methods=['PUT'])
@login_required
def marcar_todas_notificacoes_lidas():
    try:
        if session.get('user_type') != 'funcionario':
            return jsonify({"mensagem": "Nada a marcar"}), 200

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE notificacoes SET LIDA = 1 WHERE ID_FUN = %s AND LIDA = 0", (session.get('user_id'),))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Notificações marcadas como lidas"}), 200
    except Exception:
        app.logger.exception("Erro ao marcar notificações como lidas")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


# ============================================
# ADM - TODAS AS ROTAS
# ============================================

@app.route('/api/adm', methods=['POST'])
@adm_required
def cadastrar_adm():
    try:
        dados = request.get_json()

        if not dados.get('nome') or not dados.get('email') or not dados.get('cpf'):
            return jsonify({"erro": "Nome, email e CPF são obrigatórios"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        senha_hash = generate_password_hash(dados.get('senha') or '123456')

        sql = """
        INSERT INTO adm (NOME_ADM, EMAIL_ADM, CPF_ADM, TELEFONE_ADM, CARGO_ADM, SETOR_ADM, WHATSAPP_ADM, SENHA_ADM)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.execute(sql, (
            dados['nome'], dados['email'], dados['cpf'],
            dados.get('telefone', ''), dados.get('cargo', ''),
            dados.get('setor', ''), dados.get('whatsapp', ''), senha_hash
        ))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"mensagem": "ADM cadastrado com sucesso!"}), 201
    except Exception as e:
        app.logger.exception("Erro ao cadastrar administrador")
        return jsonify({"erro": "Não foi possível cadastrar o administrador agora. Tente novamente em instantes."}), 500


@app.route('/api/adm', methods=['GET'])
@adm_required
def listar_adm():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT ID_ADM, NOME_ADM, EMAIL_ADM, CPF_ADM, TELEFONE_ADM, CARGO_ADM, SETOR_ADM, WHATSAPP_ADM FROM adm")
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(resultados), 200
    except Exception as e:
        app.logger.exception("Erro ao buscar administradores")
        return jsonify({"erro": "Não foi possível carregar os dados agora. Tente novamente em instantes."}), 500


@app.route('/api/adm/<int:id_adm>', methods=['GET'])
@adm_required
def listar_adm_id(id_adm):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT ID_ADM, NOME_ADM, EMAIL_ADM, CPF_ADM, TELEFONE_ADM, CARGO_ADM, SETOR_ADM, WHATSAPP_ADM FROM adm WHERE ID_ADM = %s", (id_adm,))
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        if resultado:
            return jsonify(resultado), 200
        return jsonify({"erro": "ADM não encontrado"}), 404
    except Exception as e:
        app.logger.exception("Erro ao buscar administrador")
        return jsonify({"erro": "Não foi possível carregar os dados agora. Tente novamente em instantes."}), 500


@app.route('/api/adm/<int:id_adm>', methods=['PUT'])
@adm_required
def atualizar_adm(id_adm):
    try:
        dados = request.get_json(force=True)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM adm WHERE ID_ADM = %s", (id_adm,))
        existente = cursor.fetchone()
        if not existente:
            cursor.close()
            conn.close()
            return jsonify({"erro": "ADM não encontrado"}), 404

        senha_final = generate_password_hash(dados['senha']) if dados.get('senha') else existente['SENHA_ADM']

        sql = """
        UPDATE adm SET NOME_ADM=%s, EMAIL_ADM=%s, CPF_ADM=%s, TELEFONE_ADM=%s,
            CARGO_ADM=%s, SETOR_ADM=%s, WHATSAPP_ADM=%s, SENHA_ADM=%s
        WHERE ID_ADM = %s
        """
        cursor.execute(sql, (
            dados.get('nome', existente['NOME_ADM']),
            dados.get('email', existente['EMAIL_ADM']),
            dados.get('cpf', existente['CPF_ADM']),
            dados.get('telefone', existente['TELEFONE_ADM']),
            dados.get('cargo', existente['CARGO_ADM']),
            dados.get('setor', existente['SETOR_ADM']),
            dados.get('whatsapp', existente['WHATSAPP_ADM']),
            senha_final, id_adm
        ))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "ADM atualizado com sucesso!"}), 200
    except Exception as e:
        app.logger.exception("Erro ao atualizar administrador")
        return jsonify({"erro": "Não foi possível salvar as alterações agora. Verifique os dados e tente novamente."}), 500


@app.route('/api/adm/<int:id_adm>', methods=['DELETE'])
@adm_required
def deletar_adm(id_adm):
    conn = None
    try:
        if id_adm == session.get('user_id'):
            return jsonify({"erro": "Você não pode excluir a própria conta de administrador enquanto está logado nela."}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM adm WHERE ID_ADM = %s", (id_adm,))
        administrador = cursor.fetchone()
        if not administrador:
            cursor.close()
            conn.close()
            return jsonify({"erro": "ADM não encontrado"}), 404

        mover_para_lixeira(cursor, 'adm', id_adm, administrador, session.get('user_name'))
        cursor.execute("DELETE FROM adm WHERE ID_ADM = %s", (id_adm,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "ADM deletado com sucesso!"}), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        if conn and conn.is_connected():
            conn.close()
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


# ============================================
# FUNCIONÁRIOS - TODAS AS ROTAS
# ============================================

@app.route('/api/funcionarios', methods=['POST'])
@adm_required
def cadastrar_fun():
    try:
        dados = request.get_json()
        if not dados.get('nome') or not dados.get('email') or not dados.get('cpf'):
            return jsonify({"erro": "Nome, email e CPF são obrigatórios"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        senha_hash = generate_password_hash(dados.get('senha') or '123456')

        sql = """
        INSERT INTO funcionarios (NOME_FUN, EMAIL_FUN, CPF_FUN, TELEFONE_FUN, CARGO_FUN, SETOR_FUN, WHATSAPP_FUN, SENHA_FUN)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (
            dados['nome'], dados['email'], dados['cpf'],
            dados.get('telefone', ''), dados.get('cargo', ''),
            dados.get('setor', ''), dados.get('whatsapp', ''), senha_hash
        ))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Funcionário cadastrado com sucesso!"}), 201
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/funcionarios', methods=['GET'])
@login_required
def listar_funcionarios():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT ID_FUN, NOME_FUN, EMAIL_FUN, CPF_FUN, TELEFONE_FUN, CARGO_FUN, SETOR_FUN, WHATSAPP_FUN FROM funcionarios")
        resultados = cursor.fetchall()

        # ---- Calcula a "bolinha" de status (verde/amarelo/vermelho) de cada funcionário
        # com base na situação dos treinamentos que ELE JÁ REALIZOU (não em relação a
        # todos os treinamentos que já existiram no sistema, mesmo os que nunca foram
        # feitos por ele): vermelho = tem ao menos um treinamento vencido; amarelo = não
        # tem nada vencido, mas tem algo vencendo nos próximos 30 dias; verde = tudo que
        # ele possui está dentro da validade; sem_dados = nunca realizou nenhum treinamento.
        cursor.execute("""
            SELECT r.ID_FUN, r.ID_TREIN, r.DATA_VENCIMENTO
            FROM relatorio r
            INNER JOIN (
                SELECT ID_FUN, ID_TREIN, MAX(DATA_VENCIMENTO) AS ULTIMO_VENCIMENTO
                FROM relatorio
                GROUP BY ID_FUN, ID_TREIN
            ) u ON u.ID_FUN = r.ID_FUN AND u.ID_TREIN = r.ID_TREIN AND u.ULTIMO_VENCIMENTO = r.DATA_VENCIMENTO
        """)
        ultimos_registros = cursor.fetchall()
        cursor.close()
        conn.close()

        registros_por_fun = {}
        for reg in ultimos_registros:
            registros_por_fun.setdefault(reg['ID_FUN'], []).append(reg)

        hoje = date.today()
        limite_30_dias = hoje + timedelta(days=30)

        for f in resultados:
            regs = registros_por_fun.get(f['ID_FUN'], [])
            vencido = any(r['DATA_VENCIMENTO'] and r['DATA_VENCIMENTO'] < hoje for r in regs)
            vencendo = any(r['DATA_VENCIMENTO'] and hoje <= r['DATA_VENCIMENTO'] <= limite_30_dias for r in regs)

            if not regs:
                f['STATUS_GERAL'] = 'sem_dados'
            elif vencido:
                f['STATUS_GERAL'] = 'vermelho'
            elif vencendo:
                f['STATUS_GERAL'] = 'amarelo'
            else:
                f['STATUS_GERAL'] = 'verde'

        return jsonify(resultados), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/funcionarios/<int:id_fun>', methods=['GET'])
@login_required
def listar_funcionario_id(id_fun):
    try:
        # funcionário comum só pode ver o próprio registro
        if session.get('user_type') == 'funcionario' and session.get('user_id') != id_fun:
            return jsonify({"erro": "Acesso negado"}), 403

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT ID_FUN, NOME_FUN, EMAIL_FUN, CPF_FUN, TELEFONE_FUN, CARGO_FUN, SETOR_FUN, WHATSAPP_FUN FROM funcionarios WHERE ID_FUN = %s", (id_fun,))
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        if resultado:
            return jsonify(resultado), 200
        return jsonify({"erro": "Funcionário não encontrado"}), 404
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/funcionarios/<int:id_fun>', methods=['PUT'])
@adm_required
def atualizar_funcionario(id_fun):
    try:
        dados = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM funcionarios WHERE ID_FUN = %s", (id_fun,))
        existente = cursor.fetchone()
        if not existente:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Funcionário não encontrado"}), 404

        senha_final = generate_password_hash(dados['senha']) if dados.get('senha') else existente['SENHA_FUN']

        sql = """
        UPDATE funcionarios SET NOME_FUN=%s, EMAIL_FUN=%s, CPF_FUN=%s, TELEFONE_FUN=%s,
            CARGO_FUN=%s, SETOR_FUN=%s, WHATSAPP_FUN=%s, SENHA_FUN=%s
        WHERE ID_FUN = %s
        """
        cursor.execute(sql, (
            dados.get('nome', existente['NOME_FUN']),
            dados.get('email', existente['EMAIL_FUN']),
            dados.get('cpf', existente['CPF_FUN']),
            dados.get('telefone', existente['TELEFONE_FUN']),
            dados.get('cargo', existente['CARGO_FUN']),
            dados.get('setor', existente['SETOR_FUN']),
            dados.get('whatsapp', existente['WHATSAPP_FUN']),
            senha_final, id_fun
        ))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Funcionário atualizado com sucesso!"}), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/funcionarios/<int:id_fun>', methods=['DELETE'])
@adm_required
def deletar_funcionario(id_fun):
    conn = None
    try:
        dados = request.get_json(silent=True) or {}
        forcar = dados.get('forcar', False)

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM funcionarios WHERE ID_FUN = %s", (id_fun,))
        funcionario = cursor.fetchone()
        if not funcionario:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Funcionário não encontrado"}), 404

        if forcar:
            cursor.execute("SELECT * FROM relatorio WHERE ID_FUN = %s", (id_fun,))
            registros_vinculados = cursor.fetchall()
            for reg in registros_vinculados:
                mover_para_lixeira(cursor, 'relatorio', reg['ID_REL'], reg, session.get('user_name'))
            cursor.execute("DELETE FROM relatorio WHERE ID_FUN = %s", (id_fun,))

        mover_para_lixeira(cursor, 'funcionarios', id_fun, funcionario, session.get('user_name'))
        cursor.execute("DELETE FROM funcionarios WHERE ID_FUN = %s", (id_fun,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Funcionário deletado com sucesso!"}), 200

    except mysql.connector.errors.IntegrityError:
        if conn and conn.is_connected():
            conn.close()
        return jsonify({
            "erro": "vinculado",
            "mensagem": "Este funcionário possui treinamentos registrados no histórico. Deseja excluir o funcionário e todo o histórico de treinamentos dele?"
        }), 409
    except Exception as e:
        if conn and conn.is_connected():
            conn.close()
        return jsonify({"erro": str(e)}), 500


# ============================================
# TREINAMENTOS - TODAS AS ROTAS
# ============================================

@app.route('/api/treinamentos', methods=['POST'])
@adm_required
def cadastrar_trein():
    try:
        dados = request.get_json()
        if not dados.get('nome'):
            return jsonify({"erro": "Nome do treinamento é obrigatório"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO treinamentos (NOME_NR, VALIDADE_DIAS, DESCRICAO_NR) VALUES (%s, %s, %s)",
                       (dados['nome'], dados.get('validade', 0), dados.get('descricao', '')))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Treinamento cadastrado com sucesso!"}), 201
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/treinamentos', methods=['GET'])
@login_required
def listar_treinamentos():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM treinamentos")
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(resultados), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/treinamentos/<int:id_trein>', methods=['GET'])
@login_required
def listar_treinamento_id(id_trein):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM treinamentos WHERE ID_TREIN = %s", (id_trein,))
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        if resultado:
            return jsonify(resultado), 200
        return jsonify({"erro": "Treinamento não encontrado"}), 404
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/treinamentos/<int:id_trein>', methods=['PUT'])
@adm_required
def atualizar_treinamento(id_trein):
    try:
        dados = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM treinamentos WHERE ID_TREIN = %s", (id_trein,))
        existente = cursor.fetchone()
        if not existente:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Treinamento não encontrado"}), 404

        nova_validade = dados.get('validade', existente['VALIDADE_DIAS'])

        cursor.execute("UPDATE treinamentos SET NOME_NR=%s, VALIDADE_DIAS=%s, DESCRICAO_NR=%s WHERE ID_TREIN=%s", (
            dados.get('nome', existente['NOME_NR']),
            nova_validade,
            dados.get('descricao', existente.get('DESCRICAO_NR', '')),
            id_trein
        ))

        # Quando a validade (em dias) do treinamento muda, os relatórios já
        # gerados para ele ficam desatualizados (a DATA_VENCIMENTO foi
        # calculada com a validade antiga). Recalculamos aqui a DATA_VENCIMENTO
        # de todo relatório vinculado a este treinamento, a partir da respectiva
        # DATA_REALIZACAO + nova validade em dias, para manter os relatórios
        # (e os status Válido/Vencendo/Vencido derivados deles) consistentes
        # com o treinamento editado.
        cursor.execute(
            "UPDATE relatorio SET DATA_VENCIMENTO = DATE_ADD(DATA_REALIZACAO, INTERVAL %s DAY) "
            "WHERE ID_TREIN = %s AND DATA_REALIZACAO IS NOT NULL",
            (nova_validade, id_trein)
        )

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Treinamento atualizado com sucesso!"}), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/treinamentos/<int:id_trein>', methods=['DELETE'])
@adm_required
def deletar_treinamentos(id_trein):
    conn = None
    try:
        dados = request.get_json(silent=True) or {}
        forcar = dados.get('forcar', False)

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM treinamentos WHERE ID_TREIN = %s", (id_trein,))
        treinamento = cursor.fetchone()
        if not treinamento:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Treinamento não encontrado"}), 404

        if forcar:
            cursor.execute("SELECT * FROM relatorio WHERE ID_TREIN = %s", (id_trein,))
            registros_vinculados = cursor.fetchall()
            for reg in registros_vinculados:
                mover_para_lixeira(cursor, 'relatorio', reg['ID_REL'], reg, session.get('user_name'))
            cursor.execute("DELETE FROM relatorio WHERE ID_TREIN = %s", (id_trein,))

        mover_para_lixeira(cursor, 'treinamentos', id_trein, treinamento, session.get('user_name'))
        cursor.execute("DELETE FROM treinamentos WHERE ID_TREIN = %s", (id_trein,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Treinamento deletado com sucesso!"}), 200

    except mysql.connector.errors.IntegrityError:
        if conn and conn.is_connected():
            conn.close()
        return jsonify({
            "erro": "vinculado",
            "mensagem": "Este treinamento possui registros de realização no histórico. Deseja excluir o treinamento e todo o histórico vinculado a ele?"
        }), 409
    except Exception as e:
        if conn and conn.is_connected():
            conn.close()
        return jsonify({"erro": str(e)}), 500


# ============================================
# RELATÓRIOS - TODAS AS ROTAS
# ============================================

@app.route('/api/relatorio', methods=['POST'])
@adm_required
def cadastrar_rel():
    try:
        dados = request.get_json()
        if not dados.get('id_fun') or not dados.get('id_trein'):
            return jsonify({"erro": "ID do funcionário e ID do treinamento são obrigatórios"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
        INSERT INTO relatorio (ID_FUN, ID_TREIN, DATA_REALIZACAO, DATA_VENCIMENTO)
        VALUES (%s, %s, %s, %s)
        """, (dados['id_fun'], dados['id_trein'], dados.get('data_realizacao'), dados.get('data_vencimento')))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Relatório cadastrado com sucesso!"}), 201
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/relatorio/<int:id_rel>', methods=['DELETE'])
@adm_required
def deletar_relatorio(id_rel):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM relatorio WHERE ID_REL = %s", (id_rel,))
        relatorio = cursor.fetchone()
        if not relatorio:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Relatório não encontrado"}), 404

        mover_para_lixeira(cursor, 'relatorio', id_rel, relatorio, session.get('user_name'))
        cursor.execute("DELETE FROM relatorio WHERE ID_REL = %s", (id_rel,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Relatório excluído com sucesso!"}), 200
    except Exception as e:
        if conn and conn.is_connected():
            conn.close()
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/relatorio', methods=['GET'])
@login_required
def listar_relatorios():
    try:
        # funcionário comum só enxerga os próprios relatórios
        id_fun_filtro = request.args.get('id_fun')
        if session.get('user_type') == 'funcionario':
            id_fun_filtro = session.get('user_id')

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        sql = """
        SELECT r.*, f.NOME_FUN, t.NOME_NR
        FROM relatorio r
        LEFT JOIN funcionarios f ON r.ID_FUN = f.ID_FUN
        LEFT JOIN treinamentos t ON r.ID_TREIN = t.ID_TREIN
        """
        params = ()
        if id_fun_filtro:
            sql += " WHERE r.ID_FUN = %s"
            params = (id_fun_filtro,)

        cursor.execute(sql, params)
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()
        # IMPORTANTE: o mysql-connector devolve DATA_REALIZACAO/DATA_VENCIMENTO
        # como objetos `date` do Python. O conversor JSON padrao do Flask
        # serializa esses objetos no formato RFC 1123 (ex.: "Sat, 25 Jul 2026
        # 00:00:00 GMT"), e NAO em "AAAA-MM-DD". O front-end (status-utils.js)
        # espera "AAAA-MM-DD" e faz comparacao de string para calcular o status
        # (Valido/Vencendo/Vencido); com o formato errado, a comparacao de
        # string sempre dava "maior" (letras > digitos), fazendo todo
        # treinamento aparecer como "Valido" mesmo vencido/vencendo.
        # Convertendo aqui para ISO ("AAAA-MM-DD") corrige o calculo de status
        # em Relatorios, Treinamentos e no Dashboard, que consomem esta rota.
        for r in resultados:
            if r.get("DATA_REALIZACAO"):
                r["DATA_REALIZACAO"] = r["DATA_REALIZACAO"].isoformat()
            if r.get("DATA_VENCIMENTO"):
                r["DATA_VENCIMENTO"] = r["DATA_VENCIMENTO"].isoformat()

        return jsonify(resultados), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


def obter_registros_relatorio_filtrados():
    """
    FUNÇÃO AUXILIAR COMPARTILHADA PELAS 3 EXPORTAÇÕES (PDF, XLSX e CSV).
    Busca no banco os registros de `relatorio` (com o nome do funcionário e da
    norma já "juntados" via LEFT JOIN) e aplica os mesmos filtros de texto/status
    usados na tela de Relatórios (querystring ?termo=...&status=...).

    Regra de visibilidade:
    - ADM: vê os registros de todos os funcionários.
    - Funcionário: só vê os próprios registros (mesma regra de /api/relatorio).

    Retorna uma tupla (tipo_usuario, lista_de_registros), onde cada registro é um
    dicionário com uma chave extra '_status' (valido/vencendo/vencido).
    """
    tipo = session.get('user_type')
    termo = (request.args.get('termo') or '').strip().lower()
    status_filtro = (request.args.get('status') or '').strip().lower()

    id_fun_filtro = request.args.get('id_fun')
    if tipo == 'funcionario':
        id_fun_filtro = session.get('user_id')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    sql = """
    SELECT r.*, f.NOME_FUN, t.NOME_NR
    FROM relatorio r
    LEFT JOIN funcionarios f ON r.ID_FUN = f.ID_FUN
    LEFT JOIN treinamentos t ON r.ID_TREIN = t.ID_TREIN
    """
    params = ()
    if id_fun_filtro:
        sql += " WHERE r.ID_FUN = %s"
        params = (id_fun_filtro,)
    sql += " ORDER BY r.DATA_VENCIMENTO ASC"
    cursor.execute(sql, params)
    registros = cursor.fetchall()
    cursor.close()
    conn.close()

    # Aplica os mesmos filtros de texto/status usados na tela de Relatórios.
    registros_filtrados = []
    for r in registros:
        st = status_relatorio(r.get('DATA_VENCIMENTO'))
        if status_filtro and st != status_filtro:
            continue
        if termo:
            alvo = f"{r.get('NOME_FUN') or ''} {r.get('NOME_NR') or ''}".lower()
            if termo not in alvo:
                continue
        r['_status'] = st
        registros_filtrados.append(r)

    return tipo, registros_filtrados


@app.route('/api/relatorio/pdf', methods=['GET'])
@login_required
def exportar_relatorio_pdf():
    """
    Gera um PDF do relatório de treinamentos.
    - ADM: vê todos os funcionários (pode ainda filtrar por texto/status via querystring,
      espelhando os filtros aplicados na tela).
    - Funcionário: só enxerga os próprios registros (mesma regra de /api/relatorio).
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        tipo, registros_filtrados = obter_registros_relatorio_filtrados()

        total = len(registros_filtrados)
        validos = sum(1 for r in registros_filtrados if r['_status'] == 'valido')
        vencendo = sum(1 for r in registros_filtrados if r['_status'] == 'vencendo')
        vencidos = sum(1 for r in registros_filtrados if r['_status'] == 'vencido')

        texto_status = {'valido': 'Válido', 'vencendo': 'Vencendo', 'vencido': 'Vencido', '': '-'}
        cor_status = {
            'Válido': colors.HexColor('#16a34a'),
            'Vencendo': colors.HexColor('#d97706'),
            'Vencido': colors.HexColor('#dc2626'),
        }

        # ---------- Monta o PDF ----------
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm
        )
        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle(
            'TituloRelatorio', parent=styles['Title'], fontSize=18,
            textColor=colors.HexColor('#4c1d95'), spaceAfter=2
        )
        subtitulo_style = ParagraphStyle(
            'Subtitulo', parent=styles['Normal'], fontSize=10,
            textColor=colors.HexColor('#6b7280'), spaceAfter=14
        )

        story = [Paragraph('Relatório de Treinamentos NR', titulo_style)]
        subtitulo = ('Situação dos treinamentos realizados por todos os funcionários'
                     if tipo == 'adm' else 'Situação dos seus treinamentos realizados')
        story.append(Paragraph(
            f"{subtitulo} — Gerado em {date.today().strftime('%d/%m/%Y')} por {session.get('user_name') or ''}",
            subtitulo_style
        ))

        # Cartões de resumo (mesmos números exibidos na tela).
        resumo_dados = [
            ['Total', 'Válidos', 'Vencendo em 30 dias', 'Vencidos'],
            [str(total), str(validos), str(vencendo), str(vencidos)],
        ]
        resumo_tabela = Table(resumo_dados, colWidths=[doc.width / 4] * 4)
        resumo_tabela.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ede9fe')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4c1d95')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ddd6fe')),
        ]))
        story.append(resumo_tabela)
        story.append(Spacer(1, 16))

        cabecalho = (['Funcionário'] if tipo == 'adm' else []) + ['Treinamento', 'Realizado em', 'Vencimento', 'Status']
        linhas = [cabecalho]
        for r in registros_filtrados:
            linha = []
            if tipo == 'adm':
                linha.append(r.get('NOME_FUN') or '-')
            linha.append(r.get('NOME_NR') or '-')
            linha.append(r['DATA_REALIZACAO'].strftime('%d/%m/%Y') if r.get('DATA_REALIZACAO') else '-')
            linha.append(r['DATA_VENCIMENTO'].strftime('%d/%m/%Y') if r.get('DATA_VENCIMENTO') else '-')
            linha.append(texto_status.get(r['_status'], '-'))
            linhas.append(linha)

        if len(linhas) == 1:
            linhas.append(['Nenhum registro encontrado.'] + [''] * (len(cabecalho) - 1))

        if tipo == 'adm':
            larguras = [doc.width * 0.28] + [doc.width * 0.18] * 4
        else:
            larguras = [doc.width * 0.34, doc.width * 0.22, doc.width * 0.22, doc.width * 0.22]

        tabela = Table(linhas, colWidths=larguras, repeatRows=1)
        estilo_tabela = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6d28d9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]
        # Colore a coluna de status conforme o valor (Válido/Vencendo/Vencido).
        for i, linha in enumerate(linhas[1:], start=1):
            if linha[-1] in cor_status:
                estilo_tabela.append(('TEXTCOLOR', (-1, i), (-1, i), cor_status[linha[-1]]))
                estilo_tabela.append(('FONTNAME', (-1, i), (-1, i), 'Helvetica-Bold'))
        tabela.setStyle(TableStyle(estilo_tabela))
        story.append(tabela)

        doc.build(story)
        buffer.seek(0)

        nome_arquivo = f"relatorio_treinamentos_{date.today().strftime('%Y%m%d')}.pdf"
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=nome_arquivo)

    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/relatorio/xlsx', methods=['GET'])
@login_required
def exportar_relatorio_xlsx():
    """
    Gera uma planilha Excel (.xlsx) do relatório de treinamentos, usando os
    mesmos dados e filtros (termo/status) da exportação em PDF.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        tipo, registros_filtrados = obter_registros_relatorio_filtrados()
        texto_status = {'valido': 'Válido', 'vencendo': 'Vencendo', 'vencido': 'Vencido', '': '-'}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Relatório'

        # Cabeçalho da planilha: mostra "Funcionário" só para o ADM (funcionário
        # comum já sabe que os dados são dele, então essa coluna é omitida).
        cabecalho = (['Funcionário'] if tipo == 'adm' else []) + ['Treinamento', 'Realizado em', 'Vencimento', 'Status']
        ws.append(cabecalho)

        # Estiliza a linha de cabeçalho (fundo roxo, texto branco em negrito),
        # seguindo a mesma paleta usada no restante do sistema.
        fundo_roxo = PatternFill(start_color='6D28D9', end_color='6D28D9', fill_type='solid')
        for col_idx in range(1, len(cabecalho) + 1):
            celula = ws.cell(row=1, column=col_idx)
            celula.font = Font(bold=True, color='FFFFFF')
            celula.fill = fundo_roxo
            celula.alignment = Alignment(horizontal='center')

        # Preenche uma linha por registro, na mesma ordem de colunas do cabeçalho.
        for r in registros_filtrados:
            linha = []
            if tipo == 'adm':
                linha.append(r.get('NOME_FUN') or '-')
            linha.append(r.get('NOME_NR') or '-')
            linha.append(r['DATA_REALIZACAO'].strftime('%d/%m/%Y') if r.get('DATA_REALIZACAO') else '-')
            linha.append(r['DATA_VENCIMENTO'].strftime('%d/%m/%Y') if r.get('DATA_VENCIMENTO') else '-')
            linha.append(texto_status.get(r['_status'], '-'))
            ws.append(linha)

        # Ajusta a largura das colunas de acordo com o maior conteúdo de cada uma.
        for col_idx, titulo_coluna in enumerate(cabecalho, start=1):
            maior_tamanho = max(
                [len(titulo_coluna)] + [len(str(ws.cell(row=i, column=col_idx).value or '')) for i in range(2, ws.max_row + 1)]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = maior_tamanho + 4

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nome_arquivo = f"relatorio_treinamentos_{date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/relatorio/csv', methods=['GET'])
@login_required
def exportar_relatorio_csv():
    """
    Gera um arquivo CSV do relatório de treinamentos, usando os mesmos dados e
    filtros (termo/status) da exportação em PDF/XLSX.
    """
    try:
        import csv
        import io

        tipo, registros_filtrados = obter_registros_relatorio_filtrados()
        texto_status = {'valido': 'Válido', 'vencendo': 'Vencendo', 'vencido': 'Vencido', '': '-'}

        # Monta o CSV em memória como texto e, no fim, converte para bytes.
        # Usamos ';' como separador (padrão do Excel em português) e UTF-8 com
        # BOM (utf-8-sig) para que acentos apareçam corretos ao abrir no Excel.
        buffer_texto = io.StringIO()
        escritor = csv.writer(buffer_texto, delimiter=';')

        cabecalho = (['Funcionário'] if tipo == 'adm' else []) + ['Treinamento', 'Realizado em', 'Vencimento', 'Status']
        escritor.writerow(cabecalho)

        for r in registros_filtrados:
            linha = []
            if tipo == 'adm':
                linha.append(r.get('NOME_FUN') or '-')
            linha.append(r.get('NOME_NR') or '-')
            linha.append(r['DATA_REALIZACAO'].strftime('%d/%m/%Y') if r.get('DATA_REALIZACAO') else '-')
            linha.append(r['DATA_VENCIMENTO'].strftime('%d/%m/%Y') if r.get('DATA_VENCIMENTO') else '-')
            linha.append(texto_status.get(r['_status'], '-'))
            escritor.writerow(linha)

        buffer_bytes = BytesIO(buffer_texto.getvalue().encode('utf-8-sig'))

        nome_arquivo = f"relatorio_treinamentos_{date.today().strftime('%Y%m%d')}.csv"
        return send_file(
            buffer_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/relatorio/<int:id_rel>', methods=['GET'])
@login_required
def listar_relatorio_id(id_rel):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
        SELECT r.*, f.NOME_FUN, t.NOME_NR
        FROM relatorio r
        LEFT JOIN funcionarios f ON r.ID_FUN = f.ID_FUN
        LEFT JOIN treinamentos t ON r.ID_TREIN = t.ID_TREIN
        WHERE r.ID_REL = %s
        """, (id_rel,))
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        if not resultado:
            return jsonify({"erro": "Relatório não encontrado"}), 404
        if session.get('user_type') == 'funcionario' and resultado['ID_FUN'] != session.get('user_id'):
            return jsonify({"erro": "Acesso negado"}), 403
        return jsonify(resultado), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/relatorio/<int:id_rel>', methods=['PUT'])
@adm_required
def atualizar_relatorio(id_rel):
    try:
        dados = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM relatorio WHERE ID_REL = %s", (id_rel,))
        existente = cursor.fetchone()
        if not existente:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Relatório não encontrado"}), 404

        cursor.execute("""
        UPDATE relatorio SET ID_FUN=%s, ID_TREIN=%s, DATA_REALIZACAO=%s, DATA_VENCIMENTO=%s
        WHERE ID_REL = %s
        """, (
            dados.get('id_fun', existente['ID_FUN']),
            dados.get('id_trein', existente['ID_TREIN']),
            dados.get('data_realizacao', existente['DATA_REALIZACAO']),
            dados.get('data_vencimento', existente['DATA_VENCIMENTO']),
            id_rel
        ))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Relatório atualizado com sucesso!"}), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


@app.route('/api/relatorio/<int:id_rel>', methods=['DELETE'])
@adm_required
def deletar_rel(id_rel):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM relatorio WHERE ID_REL = %s", (id_rel,))
        registro = cursor.fetchone()
        if not registro:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Relatório não encontrado"}), 404

        mover_para_lixeira(cursor, 'relatorio', id_rel, registro, session.get('user_name'))
        cursor.execute("DELETE FROM relatorio WHERE ID_REL = %s", (id_rel,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Relatório deletado com sucesso!"}), 200
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        if conn and conn.is_connected():
            conn.close()
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


# ============================================
# FUNCIONÁRIO REALIZA TREINAMENTO
# ============================================

@app.route('/api/funcionario/realizar-treinamento', methods=['POST'])
@login_required
def realizar_treinamento():
    try:
        # Apenas usuários logados como "funcionario" podem marcar um treinamento como realizado
        # (o ADM cadastra/edita treinamentos, mas quem "realiza" é o funcionário).
        if session.get('user_type') != 'funcionario':
            return jsonify({"erro": "Apenas funcionários podem realizar treinamentos por aqui"}), 403

        # Pega do corpo da requisição (JSON) qual treinamento o funcionário está realizando.
        dados = request.get_json()
        id_trein = dados.get('id_trein')
        if not id_trein:
            return jsonify({"erro": "Selecione um treinamento"}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Confirma que o treinamento informado realmente existe antes de prosseguir.
        cursor.execute("SELECT * FROM treinamentos WHERE ID_TREIN = %s", (id_trein,))
        treinamento = cursor.fetchone()
        if not treinamento:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Treinamento não encontrado"}), 404

        hoje = date.today()

        # Busca se já existe algum registro (linha da tabela `relatorio`) desse funcionário
        # para esse treinamento. Também trazemos o ID_REL para podermos ATUALIZAR essa linha
        # depois, em vez de criar uma nova — isso é o que evita registros duplicados quando um
        # treinamento pendente/vencido é refeito: o registro antigo é substituído pelo novo.
        cursor.execute("""
            SELECT ID_REL, DATA_VENCIMENTO FROM relatorio
            WHERE ID_FUN = %s AND ID_TREIN = %s
            ORDER BY DATA_VENCIMENTO DESC
            LIMIT 1
        """, (session.get('user_id'), id_trein))
        ultimo = cursor.fetchone()

        # Se já existe um registro e ele ainda está "válido" (faltam mais de 30 dias para
        # vencer), o funcionário não pode refazer o treinamento antes da hora — só quando
        # estiver pendente/vencendo/vencido é que a realização é permitida.
        if ultimo and ultimo['DATA_VENCIMENTO']:
            data_liberacao = ultimo['DATA_VENCIMENTO'] - timedelta(days=30)
            if hoje < data_liberacao:
                cursor.close()
                conn.close()
                return jsonify({
                    "erro": f"Você já realizou este treinamento e ele ainda está válido até "
                            f"{ultimo['DATA_VENCIMENTO'].strftime('%d/%m/%Y')}. "
                            f"Só será possível refazê-lo a partir de {data_liberacao.strftime('%d/%m/%Y')} "
                            f"(30 dias antes do vencimento)."
                }), 409

        # Calcula a nova data de vencimento com base na validade (em dias) do treinamento.
        vencimento = hoje + timedelta(days=treinamento['VALIDADE_DIAS'])

        if ultimo:
            # Já existia um registro pendente/vencido desse treinamento para o funcionário:
            # em vez de inserir uma nova linha (o que geraria duplicidade no histórico),
            # ATUALIZAMOS o registro existente, substituindo a data de realização e o
            # vencimento antigos pelos novos. Ou seja, o treinamento "pendente" é substituído
            # pelo treinamento "realizado".
            cursor.execute("""
                UPDATE relatorio
                SET DATA_REALIZACAO = %s, DATA_VENCIMENTO = %s
                WHERE ID_REL = %s
            """, (hoje, vencimento, ultimo['ID_REL']))
        else:
            # É a primeira vez que o funcionário realiza esse treinamento: como não existe
            # registro anterior para substituir, criamos uma linha nova em `relatorio`.
            cursor.execute("""
            INSERT INTO relatorio (ID_FUN, ID_TREIN, DATA_REALIZACAO, DATA_VENCIMENTO)
            VALUES (%s, %s, %s, %s)
            """, (session.get('user_id'), id_trein, hoje, vencimento))

        # Confirma (salva) as alterações no banco e libera a conexão.
        conn.commit()
        cursor.close()
        conn.close()

        vencimento_br = vencimento.strftime('%d/%m/%Y')
        return jsonify({"mensagem": f"Treinamento {treinamento['NOME_NR']} realizado com sucesso! Válido até {vencimento_br}."}), 201
    except Exception as e:
        app.logger.exception("Erro inesperado no servidor")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


# ============================================
# LIXEIRA - RECUPERAÇÃO DE DADOS EXCLUÍDOS
# ============================================

# Rótulos amigáveis para exibir de onde veio cada item excluído.
NOME_TABELA_ORIGEM = {
    'funcionarios': 'Funcionário',
    'adm': 'Administrador',
    'treinamentos': 'Treinamento (NR)',
    'relatorio': 'Registro de treinamento realizado',
}


@app.route('/api/lixeira', methods=['GET'])
@adm_required
def listar_lixeira():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM lixeira ORDER BY DATA_EXCLUSAO DESC LIMIT 300")
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()

        for r in resultados:
            if r.get('DATA_EXCLUSAO'):
                r['DATA_EXCLUSAO'] = r['DATA_EXCLUSAO'].isoformat()
            # O conector pode devolver a coluna JSON já como dict ou como texto,
            # dependendo da versão; normalizamos para dict aqui.
            if isinstance(r.get('DADOS'), str):
                try:
                    r['DADOS'] = json.loads(r['DADOS'])
                except Exception:
                    pass
            r['TABELA_ORIGEM_LABEL'] = NOME_TABELA_ORIGEM.get(r['TABELA_ORIGEM'], r['TABELA_ORIGEM'])

        return jsonify(resultados), 200
    except Exception:
        app.logger.exception("Erro ao listar itens da lixeira")
        return jsonify({"erro": "Não foi possível carregar os itens excluídos agora. Tente novamente em instantes."}), 500


@app.route('/api/lixeira/<int:id_lixeira>/restaurar', methods=['POST'])
@adm_required
def restaurar_lixeira(id_lixeira):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM lixeira WHERE ID_LIXEIRA = %s", (id_lixeira,))
        item = cursor.fetchone()
        if not item:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Item não encontrado na lixeira"}), 404

        tabela = item['TABELA_ORIGEM']
        colunas = TABELAS_RESTAURAVEIS.get(tabela)
        if not colunas:
            cursor.close()
            conn.close()
            return jsonify({"erro": "Não é possível restaurar este tipo de registro"}), 400

        dados = item['DADOS']
        if isinstance(dados, str):
            dados = json.loads(dados)

        # relatorio referencia funcionarios/treinamentos: se o funcionário ou o
        # treinamento também tiverem sido excluídos (e não restaurados antes),
        # o INSERT abaixo falha por causa da FOREIGN KEY. Nesse caso avisamos
        # o usuário em vez de deixar um erro genérico estourar.
        if tabela == 'relatorio':
            cursor.execute("SELECT 1 FROM funcionarios WHERE ID_FUN = %s", (dados.get('ID_FUN'),))
            fun_existe = cursor.fetchone()
            cursor.execute("SELECT 1 FROM treinamentos WHERE ID_TREIN = %s", (dados.get('ID_TREIN'),))
            trein_existe = cursor.fetchone()
            if not fun_existe or not trein_existe:
                cursor.close()
                conn.close()
                return jsonify({
                    "erro": "Para restaurar este registro é preciso primeiro restaurar o funcionário e/ou o treinamento vinculados a ele."
                }), 409

        valores = [dados.get(c) for c in colunas]
        placeholders = ', '.join(['%s'] * len(colunas))
        colunas_sql = ', '.join(f'`{c}`' for c in colunas)

        cursor_escrita = conn.cursor()
        try:
            cursor_escrita.execute(
                f"INSERT INTO `{tabela}` ({colunas_sql}) VALUES ({placeholders})",
                valores
            )
        except mysql.connector.errors.IntegrityError:
            cursor.close()
            cursor_escrita.close()
            conn.close()
            return jsonify({"erro": "Já existe um registro com os mesmos dados (e-mail, CPF ou telefone). Não foi possível restaurar."}), 409

        cursor_escrita.execute("DELETE FROM lixeira WHERE ID_LIXEIRA = %s", (id_lixeira,))
        conn.commit()
        cursor.close()
        cursor_escrita.close()
        conn.close()
        return jsonify({"mensagem": "Registro restaurado com sucesso!"}), 200
    except Exception:
        app.logger.exception("Erro ao restaurar item da lixeira")
        if conn and conn.is_connected():
            conn.close()
        return jsonify({"erro": "Não foi possível restaurar este item agora. Tente novamente em instantes."}), 500


@app.route('/api/lixeira/<int:id_lixeira>', methods=['DELETE'])
@adm_required
def excluir_definitivo_lixeira(id_lixeira):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM lixeira WHERE ID_LIXEIRA = %s", (id_lixeira,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"mensagem": "Item excluído definitivamente."}), 200
    except Exception:
        app.logger.exception("Erro ao excluir item da lixeira definitivamente")
        return jsonify({"erro": "Não foi possível concluir a operação agora. Tente novamente em instantes."}), 500


if __name__ == '__main__':
    app.run(debug=True)
